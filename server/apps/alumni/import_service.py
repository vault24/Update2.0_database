"""
Alumni bulk-import service.

Reads a tabular source (Excel / CSV / Google Sheet), maps arbitrary column
headers onto our fields via `import_config`, validates, and creates the
Student + Alumni pair for every row inside one transaction.

Design notes
------------
* **One import writes two tables.** A spreadsheet routinely carries columns
  that live on the Student record (father's name, address, board roll…) rather
  than on Alumni. Those are never discarded: `import_config` tags each field
  with its target table, and the row payload is handed to the same
  `create_alumni_from_essentials` contract the manual-add flow uses, which
  populates Student and Alumni together.
* **Validate everything before writing anything.** Each row is checked and
  reported on individually (so admins get row-wise errors), and only the rows
  that survive are written — via `bulk_create`, in a single transaction. A
  bad row therefore never aborts a good one, while a database-level failure
  rolls the whole batch back.
* **Subset-friendly.** Only the fields marked required in `import_config` are
  enforced; every other column is optional, and unknown columns are ignored
  (and reported, so the admin can see what was skipped).
"""
import csv
import io
import logging
import re
import uuid
from datetime import date, datetime
from decimal import Decimal, InvalidOperation

from django.db import transaction
from django.utils import timezone

from .import_config import (
    FIELD_SPECS,
    REQUIRED_SPECS,
    TYPE_DATE,
    TYPE_DECIMAL,
    TYPE_DEPARTMENT,
    TYPE_EMAIL,
    TYPE_INT,
    TYPE_CHOICE,
    normalize_header,
    resolve_header,
)

logger = logging.getLogger(__name__)

# Guard rails for a single upload.
#
# There is deliberately no file-size cap: admins import whatever their office
# spreadsheet happens to weigh, and a byte limit is a poor proxy for cost
# anyway (a 30 MB .xlsx can hold fewer rows than a 2 MB CSV). MAX_ROWS is the
# real protection — it bounds the work regardless of how the file is encoded.
# Nginx's client_max_body_size still bounds the HTTP upload itself.
MAX_ROWS = 5000

# Only the remote Google-Sheets fetch keeps a ceiling: that download is not
# covered by Nginx's request limit, so it needs its own bound.
MAX_SHEET_FETCH_BYTES = 100 * 1024 * 1024  # 100 MB

# Google Sheets is the only host we will fetch server-side (SSRF guard).
_GOOGLE_SHEET_HOSTS = ('docs.google.com',)
_SHEET_ID_RE = re.compile(r'/spreadsheets/d/([a-zA-Z0-9-_]+)')
_GID_RE = re.compile(r'[#&?]gid=([0-9]+)')

_EMAIL_RE = re.compile(r'^[^\s@]+@[^\s@]+\.[^\s@]+$')


class ImportError_(Exception):
    """Raised for problems with the *source* (not with an individual row)."""


# ---------------------------------------------------------------------------
# Source readers
# ---------------------------------------------------------------------------
def google_sheet_csv_url(url: str) -> str:
    """
    Turn any shareable Google Sheets URL into its CSV export URL.

    Only docs.google.com is accepted: this URL is fetched by the server, so an
    open-ended fetch would be an SSRF hole (internal metadata endpoints etc.).
    """
    from urllib.parse import urlparse

    parsed = urlparse((url or '').strip())
    if parsed.scheme not in ('http', 'https'):
        raise ImportError_('The sheet link must be a full http(s) URL.')
    if parsed.hostname not in _GOOGLE_SHEET_HOSTS:
        raise ImportError_(
            'Only Google Sheets links (docs.google.com) can be imported by URL. '
            'For any other source, download the sheet and upload the file.'
        )
    match = _SHEET_ID_RE.search(parsed.path)
    if not match:
        raise ImportError_('That does not look like a Google Sheets link.')

    sheet_id = match.group(1)
    gid_match = _GID_RE.search(url)
    gid = gid_match.group(1) if gid_match else '0'
    return f'https://docs.google.com/spreadsheets/d/{sheet_id}/export?format=csv&gid={gid}'


def fetch_google_sheet(url: str) -> bytes:
    """Download a public Google Sheet as CSV bytes."""
    import urllib.request
    import urllib.error

    export_url = google_sheet_csv_url(url)
    request = urllib.request.Request(export_url, headers={'User-Agent': 'SIPI-Importer/1.0'})
    try:
        with urllib.request.urlopen(request, timeout=20) as response:
            final_host = response.geturl().split('/')[2] if '//' in response.geturl() else ''
            # Google redirects unshared sheets to a login page on another host.
            if final_host and not final_host.endswith('google.com'):
                raise ImportError_('The sheet redirected off Google — is the link public?')
            data = response.read(MAX_SHEET_FETCH_BYTES + 1)
    except urllib.error.HTTPError as exc:
        if exc.code in (401, 403):
            raise ImportError_(
                'Google refused access to that sheet. Set its sharing to '
                '"Anyone with the link can view" and try again.'
            ) from exc
        raise ImportError_(f'Could not download the sheet (HTTP {exc.code}).') from exc
    except urllib.error.URLError as exc:
        raise ImportError_(f'Could not reach Google Sheets: {exc.reason}') from exc

    if len(data) > MAX_SHEET_FETCH_BYTES:
        raise ImportError_('That sheet is too large to download.')
    if b'<html' in data[:200].lower():
        raise ImportError_(
            'Google returned a web page instead of CSV — the sheet is probably '
            'not shared publicly.'
        )
    return data


def _read_csv(data: bytes):
    for encoding in ('utf-8-sig', 'utf-8', 'cp1252', 'latin-1'):
        try:
            text = data.decode(encoding)
            break
        except UnicodeDecodeError:
            continue
    else:
        raise ImportError_('Could not decode the CSV file — please save it as UTF-8.')

    try:
        sample = text[:4096]
        dialect = csv.Sniffer().sniff(sample, delimiters=',;\t|')
    except csv.Error:
        dialect = csv.excel

    reader = csv.reader(io.StringIO(text), dialect)
    rows = [row for row in reader]
    if not rows:
        raise ImportError_('The file is empty.')
    return rows[0], rows[1:]


def _read_xlsx(data: bytes):
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover - dependency is pinned
        raise ImportError_(
            'Excel support is unavailable on the server (openpyxl is not installed).'
        ) from exc

    try:
        workbook = load_workbook(
            io.BytesIO(data), read_only=True, data_only=True, keep_links=False
        )
    except Exception as exc:  # noqa: BLE001 - openpyxl raises many shapes
        raise ImportError_(f'Could not read the Excel file: {exc}') from exc

    try:
        sheet = workbook.active
        rows = []
        for row in sheet.iter_rows(values_only=True):
            rows.append(list(row))
            if len(rows) > MAX_ROWS + 1:
                break
    finally:
        workbook.close()

    if not rows:
        raise ImportError_('The spreadsheet is empty.')
    return rows[0], rows[1:]


def read_table(*, file=None, sheet_url=None):
    """
    Read a source into (headers, rows).

    Returns:
        headers: list[str] — raw header labels as written by the admin.
        rows:    list[list] — data rows (trailing blank rows removed).
    """
    if sheet_url:
        header, rows = _read_csv(fetch_google_sheet(sheet_url))
    elif file is not None:
        data = file.read()
        if not data:
            raise ImportError_('The uploaded file is empty.')
        name = (getattr(file, 'name', '') or '').lower()
        if name.endswith('.xlsx') or data[:2] == b'PK':
            header, rows = _read_xlsx(data)
        elif name.endswith(('.csv', '.tsv', '.txt')):
            header, rows = _read_csv(data)
        elif name.endswith('.xls'):
            raise ImportError_(
                'Legacy .xls files are not supported. Open it in Excel and '
                'save as .xlsx (or export CSV).'
            )
        else:
            header, rows = _read_csv(data)
    else:
        raise ImportError_('Provide a file to upload or a Google Sheets link.')

    headers = ['' if h is None else str(h).strip() for h in header]

    # Drop rows that are entirely empty (Excel loves trailing blanks).
    cleaned = [r for r in rows if any(_is_filled(c) for c in r)]
    if len(cleaned) > MAX_ROWS:
        raise ImportError_(f'Too many rows ({len(cleaned)}). The limit is {MAX_ROWS} per import.')
    return headers, cleaned


def _is_filled(value):
    return value is not None and str(value).strip() != ''


# ---------------------------------------------------------------------------
# Column mapping
# ---------------------------------------------------------------------------
def analyze_columns(headers):
    """
    Describe how each spreadsheet column maps onto our fields.

    Returns a dict with the detected columns, which required fields were
    matched, which are missing, and any duplicate/unknown columns — everything
    the preview screen needs to let an admin decide before writing data.
    """
    columns = []
    mapped_keys = {}
    duplicates = []

    for index, header in enumerate(headers):
        spec = resolve_header(header) if header else None
        entry = {
            'index': index,
            'header': header,
            'mappedTo': spec.key if spec else None,
            'label': spec.label if spec else None,
            'target': spec.target if spec else None,
            'required': bool(spec and spec.required),
        }
        if spec:
            if spec.key in mapped_keys:
                entry['duplicate'] = True
                duplicates.append(header)
            else:
                mapped_keys[spec.key] = index
        columns.append(entry)

    matched_required = [s for s in REQUIRED_SPECS if s.key in mapped_keys]
    missing_required = [s for s in REQUIRED_SPECS if s.key not in mapped_keys]
    unknown = [c['header'] for c in columns if c['header'] and not c['mappedTo']]

    warnings = []
    if unknown:
        warnings.append(
            f"{len(unknown)} column(s) were not recognised and will be ignored: "
            f"{', '.join(unknown[:8])}{'…' if len(unknown) > 8 else ''}."
        )
    if duplicates:
        warnings.append(
            f"Duplicate column(s) detected — only the first of each is used: "
            f"{', '.join(duplicates)}."
        )

    return {
        'columns': columns,
        'mappedKeys': mapped_keys,
        'mappedRequired': [{'key': s.key, 'label': s.label} for s in matched_required],
        'missingRequired': [
            {'key': s.key, 'label': s.label, 'recommended': s.recommended}
            for s in missing_required
        ],
        'unknownColumns': unknown,
        'warnings': warnings,
        'canImport': not missing_required,
    }


# ---------------------------------------------------------------------------
# Value coercion
# ---------------------------------------------------------------------------
def _coerce(spec, raw, departments):
    """
    Convert a cell into the value our models expect.

    Returns (value, error). `value` is None when the cell is blank — callers
    simply omit the key, which is what makes every optional column optional.
    """
    if not _is_filled(raw):
        return None, None

    if isinstance(raw, str):
        raw = raw.strip()

    if spec.type == TYPE_INT:
        try:
            # Excel numeric cells arrive as floats ("2015.0").
            return int(float(str(raw).replace(',', '').strip())), None
        except (TypeError, ValueError):
            return None, f'{spec.label}: "{raw}" is not a whole number.'

    if spec.type == TYPE_DECIMAL:
        try:
            value = Decimal(str(raw).strip())
        except (InvalidOperation, TypeError, ValueError):
            return None, f'{spec.label}: "{raw}" is not a number.'
        if value < 0 or value > 5:
            return None, f'{spec.label}: {value} is out of range (0.00–5.00).'
        return value, None

    if spec.type == TYPE_DATE:
        if isinstance(raw, datetime):
            return raw.date().isoformat(), None
        if isinstance(raw, date):
            return raw.isoformat(), None
        text = str(raw).strip()
        for fmt in ('%Y-%m-%d', '%d/%m/%Y', '%d-%m-%Y', '%m/%d/%Y', '%Y/%m/%d', '%d.%m.%Y'):
            try:
                return datetime.strptime(text, fmt).date().isoformat(), None
            except ValueError:
                continue
        return None, f'{spec.label}: "{raw}" is not a recognisable date (use YYYY-MM-DD).'

    if spec.type == TYPE_EMAIL:
        text = str(raw).strip()
        if not _EMAIL_RE.match(text):
            return None, f'{spec.label}: "{raw}" is not a valid email address.'
        return text, None

    if spec.type == TYPE_CHOICE:
        text = str(raw).strip()
        norm = normalize_header(text)
        for choice in spec.choices:
            if normalize_header(choice) == norm:
                return choice, None
        # Tolerate common shorthands, e.g. "M"/"F" for gender.
        for choice in spec.choices:
            if norm and normalize_header(choice).startswith(norm):
                return choice, None
        allowed = ', '.join(spec.choices)
        return None, f'{spec.label}: "{raw}" must be one of: {allowed}.'

    if spec.type == TYPE_DEPARTMENT:
        norm = normalize_header(raw)
        department = departments.get(norm)
        if department is None:
            return None, f'{spec.label}: "{raw}" does not match any department.'
        return department, None

    return str(raw).strip(), None


def _department_lookup():
    """Map normalised department names AND codes to their Department."""
    from apps.departments.models import Department

    lookup = {}
    for department in Department.objects.all():
        for candidate in (department.code, department.name):
            norm = normalize_header(candidate)
            if norm:
                lookup.setdefault(norm, department)
    return lookup


def _assign(payload, key, value):
    """Set a possibly dotted key, building nested dicts as needed."""
    if '.' not in key:
        payload[key] = value
        return
    head, tail = key.split('.', 1)
    payload.setdefault(head, {})
    _assign(payload[head], tail, value)


# ---------------------------------------------------------------------------
# Row -> payload
# ---------------------------------------------------------------------------
def build_row_payloads(headers, rows, departments):
    """
    Convert raw rows into validated payloads.

    Returns (payloads, row_errors) where payloads is a list of
    {'rowNumber': int, 'data': dict} and row_errors is a list of
    {'rowNumber': int, 'errors': [str, ...]}.
    """
    specs_by_index = {}
    seen_keys = set()
    for index, header in enumerate(headers):
        spec = resolve_header(header) if header else None
        if spec and spec.key not in seen_keys:
            specs_by_index[index] = spec
            seen_keys.add(spec.key)

    payloads = []
    row_errors = []

    for offset, row in enumerate(rows):
        row_number = offset + 2  # +1 for the header row, +1 for 1-based rows
        payload = {}
        errors = []

        for index, spec in specs_by_index.items():
            raw = row[index] if index < len(row) else None
            value, error = _coerce(spec, raw, departments)
            if error:
                errors.append(error)
            elif value is not None:
                _assign(payload, spec.key, value)

        # Required fields must be present and non-empty.
        for spec in REQUIRED_SPECS:
            if payload.get(spec.key) in (None, ''):
                if not any(e.startswith(f'{spec.label}:') for e in errors):
                    errors.append(f'{spec.label} is required but empty.')

        if errors:
            row_errors.append({'rowNumber': row_number, 'errors': errors})
        else:
            payloads.append({'rowNumber': row_number, 'data': payload})

    return payloads, row_errors


# ---------------------------------------------------------------------------
# Import
# ---------------------------------------------------------------------------
def _gen_identifier(prefix):
    return f'{prefix}-{uuid.uuid4().hex[:10].upper()}'


def _duplicate_key(payload):
    """Identity used to skip rows already present (roll / registration)."""
    roll = (payload.get('currentRollNumber') or '').strip()
    reg = (payload.get('currentRegistrationNumber') or '').strip()
    return roll or None, reg or None


@transaction.atomic
def import_alumni(headers, rows, *, dry_run=False):
    """
    Validate and (unless dry_run) create Student+Alumni records.

    Every valid row is written with `bulk_create` inside this transaction:
    invalid rows are reported and skipped without affecting the rest, while a
    database-level failure rolls the entire batch back.

    Returns a summary dict: processed / imported / skipped / failed, plus
    row-wise errors and the mapping analysis.
    """
    from apps.students.models import Student
    from .models import Alumni

    analysis = analyze_columns(headers)
    if not analysis['canImport']:
        missing = ', '.join(m['label'] for m in analysis['missingRequired'])
        raise ImportError_(f'Required column(s) missing: {missing}.')

    departments = _department_lookup()
    payloads, row_errors = build_row_payloads(headers, rows, departments)

    # --- Duplicate detection (existing rows + within this file) -------------
    existing_rolls = set(
        Student.objects.exclude(currentRollNumber='')
        .values_list('currentRollNumber', flat=True)
    )
    existing_regs = set(
        Student.objects.exclude(currentRegistrationNumber='')
        .values_list('currentRegistrationNumber', flat=True)
    )

    accepted = []
    skipped = []
    batch_rolls, batch_regs = set(), set()

    for item in payloads:
        roll, reg = _duplicate_key(item['data'])
        reason = None
        if roll and (roll in existing_rolls or roll in batch_rolls):
            reason = f'Roll "{roll}" already exists — row skipped.'
        elif reg and (reg in existing_regs or reg in batch_regs):
            reason = f'Registration "{reg}" already exists — row skipped.'

        if reason:
            skipped.append({'rowNumber': item['rowNumber'], 'errors': [reason]})
            continue

        if roll:
            batch_rolls.add(roll)
        if reg:
            batch_regs.add(reg)
        accepted.append(item)

    summary = {
        'processed': len(rows),
        'imported': 0,
        'skipped': len(skipped),
        'failed': len(row_errors),
        'errors': sorted(row_errors + skipped, key=lambda e: e['rowNumber']),
        'analysis': analysis,
        'dryRun': dry_run,
    }

    if dry_run:
        summary['wouldImport'] = len(accepted)
        return summary

    if not accepted:
        return summary

    # --- Build objects ------------------------------------------------------
    students = []
    alumni_rows = []
    now = timezone.now()

    for item in accepted:
        data = item['data']
        present = data.get('presentAddress') or {}
        permanent = data.get('permanentAddress') or {}
        graduation_year = data.get('graduationYear')

        student = Student(
            fullNameEnglish=data['fullNameEnglish'],
            fullNameBangla=data.get('fullNameBangla', ''),
            fatherName=data.get('fatherName', ''),
            motherName=data.get('motherName', ''),
            fatherNID=data.get('fatherNID', ''),
            motherNID=data.get('motherNID', ''),
            nidNumber=data.get('nidNumber', ''),
            birthCertificateNo=data.get('birthCertificateNo', ''),
            dateOfBirth=data.get('dateOfBirth') or None,
            gender=data.get('gender', ''),
            religion=data.get('religion', ''),
            bloodGroup=data.get('bloodGroup', ''),
            nationality=data.get('nationality') or 'Bangladeshi',
            maritalStatus=data.get('maritalStatus', ''),
            mobileStudent=data.get('mobileStudent', ''),
            guardianMobile=data.get('guardianMobile', ''),
            email=data.get('email', ''),
            emergencyContact=data.get('emergencyContact', ''),
            presentAddress=present if isinstance(present, dict) else {},
            permanentAddress=permanent if isinstance(permanent, dict) else {},
            highestExam=data.get('highestExam', ''),
            board=data.get('board', ''),
            group=data.get('group', ''),
            rollNumber=data.get('rollNumber', ''),
            registrationNumber=data.get('registrationNumber', ''),
            passingYear=data.get('passingYear'),
            gpa=data.get('gpa'),
            finalCgpa=data.get('finalCgpa'),
            institutionName=data.get('institutionName', ''),
            currentRollNumber=(data.get('currentRollNumber') or _gen_identifier('ALM')),
            currentRegistrationNumber=(
                data.get('currentRegistrationNumber') or _gen_identifier('ALMREG')
            ),
            semester=8,               # graduated => completed the final semester
            department=data['department'],
            session=data.get('session', ''),
            shift=data.get('shift', ''),
            currentGroup=data.get('currentGroup', ''),
            status='graduated',
            enrollmentDate=data.get('enrollmentDate') or None,
        )
        students.append(student)

        position = data.get('currentPosition') or None
        alumni_rows.append(
            Alumni(
                student=student,
                alumniType=data.get('alumniType') or 'established',
                graduationYear=graduation_year,
                currentSupportCategory=data.get('currentSupportCategory') or 'no_support_needed',
                currentPosition=position,
                bio=data.get('bio') or None,
                linkedinUrl=data.get('linkedinUrl') or None,
                portfolioUrl=data.get('portfolioUrl') or None,
                registrationSource='admin_manual',
                reviewStatus='approved',
                isVerified=True,
                lastEditedBy='admin',
                supportHistory=[{
                    'date': now.isoformat(),
                    'previousCategory': None,
                    'newCategory': data.get('currentSupportCategory') or 'no_support_needed',
                    'notes': 'Alumni created via spreadsheet import',
                }],
            )
        )

    # Student PKs are client-side UUIDs, so the Alumni rows (whose PK *is* the
    # student FK) can be built before either table is written.
    Student.objects.bulk_create(students, batch_size=500)
    Alumni.objects.bulk_create(alumni_rows, batch_size=500)

    summary['imported'] = len(accepted)
    logger.info(
        'Alumni import: processed=%s imported=%s skipped=%s failed=%s',
        summary['processed'], summary['imported'], summary['skipped'], summary['failed'],
    )
    return summary


def preview_import(headers, rows, *, sample_size=5):
    """Analysis + validation with no writes, for the confirmation screen."""
    departments = _department_lookup()
    analysis = analyze_columns(headers)

    payloads, row_errors = ([], [])
    if analysis['canImport']:
        payloads, row_errors = build_row_payloads(headers, rows, departments)

    sample = []
    for item in payloads[:sample_size]:
        flat = {}
        for spec in FIELD_SPECS:
            if '.' in spec.key:
                head, tail = spec.key.split('.', 1)
                value = (item['data'].get(head) or {}).get(tail)
            else:
                value = item['data'].get(spec.key)
            if value not in (None, '', {}):
                flat[spec.key] = str(value)
        sample.append({'rowNumber': item['rowNumber'], 'values': flat})

    return {
        'totalRows': len(rows),
        'validRows': len(payloads),
        'invalidRows': len(row_errors),
        'sampleRows': sample,
        'errors': row_errors[:50],
        **analysis,
    }
