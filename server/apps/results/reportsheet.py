"""
Result-sheet renderers: the official BTEB-style tabulation sheet as a
colourful PDF (reportlab) or Excel workbook (openpyxl).

Both renderers consume the same structured data from
``analytics.sheet_rows`` so the two formats never drift apart. Column order
mirrors the institute's printed tabulation sheet, with the aggregate Summary
block kept in the LAST column:

    SL · Student Name · Gender · Roll · GPA · Referred · Failed ·
    Ref. Sub. Sem Wise · Position · Summary
"""
from __future__ import annotations

from io import BytesIO

from django.utils import timezone

# Palette (kept in one place so PDF + Excel share the same colours).
_HEADER_BG = '#1e3a8a'      # indigo-900
_HEADER_FG = '#ffffff'
_PASS_BG = '#dcfce7'        # emerald-100
_REFERRED_BG = '#fef3c7'    # amber-100
_FAIL_BG = '#fee2e2'        # red-100
_ALT_ROW = '#f1f5f9'        # slate-100
_SUMMARY_BG = '#e0e7ff'     # indigo-100
_BORDER = '#94a3b8'         # slate-400

INSTITUTE_NAME = 'Sirajganj Polytechnic Institute, Sirajganj'

# Column layout — shared between both renderers so PDF and Excel stay aligned.
# 0-based indices used by the PDF; +1 gives the 1-based Excel column.
HEADERS_PDF = ['SL', 'Student Name', 'Gen', 'Roll', 'GPA', 'Referred',
               'Failed', 'Ref. Sub.\nSem Wise', 'Position', 'Summary']
HEADERS_XLSX = ['SL', 'Student Name', 'Gender', 'Roll', 'GPA', 'Referred',
                'Failed', 'Ref. Sub. Sem Wise', 'Position', 'Summary']
_COL_SL, _COL_NAME, _COL_GEN, _COL_ROLL, _COL_GPA = 0, 1, 2, 3, 4
_COL_REFERRED, _COL_FAILED, _COL_REFSUB, _COL_POSITION, _COL_SUMMARY = 5, 6, 7, 8, 9


def _title(sheet: dict) -> str:
    return (
        f"Technology: {sheet['departmentName']}, "
        f"Result Semester: {sheet['semesterLabel']}, "
        f"Shift: {sheet['shift']}"
    )


def _summary_lines(summary: dict) -> list[str]:
    return [
        f"Total Student: {summary['totalStudent']}",
        f"Total Pass: {summary['totalPass']:02d}",
        f"Total Referred: {summary['totalReferred']:02d}",
        f"Total Fail: {summary['totalFail']:02d}",
        f"% Pass: {summary['pctPass']}",
        f"% Referred: {summary['pctReferred']}",
        f"% Fail: {summary['pctFail']}",
        f"Total %: {summary['pctTotal']}",
    ]


# ---------------------------------------------------------------------------
# PDF
# ---------------------------------------------------------------------------

def render_pdf(sheet: dict) -> bytes:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.units import mm
    from reportlab.platypus import (
        Paragraph,
        SimpleDocTemplate,
        Spacer,
        Table,
        TableStyle,
    )
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet

    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer, pagesize=landscape(A4),
        leftMargin=10 * mm, rightMargin=10 * mm,
        topMargin=10 * mm, bottomMargin=10 * mm,
        title=f"Result Sheet — {sheet['semesterLabel']} Semester",
    )
    styles = getSampleStyleSheet()
    cell = ParagraphStyle('cell', parent=styles['Normal'], fontSize=7, leading=8.5)
    cell_center = ParagraphStyle('cellc', parent=cell, alignment=1)
    summary_style = ParagraphStyle('summary', parent=cell, fontSize=7.5, leading=12)
    head_style = ParagraphStyle(
        'title', parent=styles['Title'], fontSize=14, spaceAfter=2,
        textColor=colors.HexColor('#1e3a8a'),
    )
    sub_style = ParagraphStyle('sub', parent=styles['Normal'], fontSize=9,
                               alignment=1, textColor=colors.HexColor('#334155'))

    story = [
        Paragraph(INSTITUTE_NAME, head_style),
        Paragraph(_title(sheet), sub_style),
        Spacer(1, 6),
    ]

    data = [HEADERS_PDF]

    summary = sheet['summary']
    # The whole summary block lives in the FIRST data row's summary cell as a
    # multi-line paragraph; the SPAN below merges the column so it reads as one
    # box (a spanned reportlab cell only renders its top-left content).
    summary_para = Paragraph('<br/>'.join(_summary_lines(summary)), summary_style)
    for index, row in enumerate(sheet['rows']):
        data.append([
            str(row['sl']),
            Paragraph(row['name'], cell),
            row['gender'],
            row['roll'],
            row['gpa'],
            Paragraph(row['referredSubjects'], cell),
            Paragraph(row['failedSubjects'], cell),
            Paragraph(row['refSubSemWise'], cell_center),
            row['position'],
            summary_para if index == 0 else '',
        ])

    # Column widths (mm) — Summary now anchors the right edge.
    #        SL Name Gen Roll GPA Ref Fail RefSub Pos Summary
    col_widths = [9, 50, 10, 20, 14, 58, 28, 22, 16, 38]
    col_widths = [w * mm for w in col_widths]

    table = Table(data, colWidths=col_widths, repeatRows=1)
    style = [
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor(_HEADER_BG)),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.HexColor(_HEADER_FG)),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 7.5),
        ('FONTSIZE', (0, 1), (-1, -1), 7),
        ('GRID', (0, 0), (-1, -1), 0.4, colors.HexColor(_BORDER)),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('ALIGN', (_COL_SL, 0), (_COL_SL, -1), 'CENTER'),
        ('ALIGN', (_COL_GEN, 0), (_COL_GPA, -1), 'CENTER'),   # Gen, Roll, GPA
        ('ALIGN', (_COL_POSITION, 0), (_COL_POSITION, -1), 'CENTER'),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('TOPPADDING', (0, 0), (-1, -1), 2),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 2),
    ]

    # Merge the Summary column into one tall block.
    n_rows = len(sheet['rows'])
    if n_rows:
        style.append(('SPAN', (_COL_SUMMARY, 1), (_COL_SUMMARY, n_rows)))
        style.append(('BACKGROUND', (_COL_SUMMARY, 1), (_COL_SUMMARY, n_rows),
                      colors.HexColor(_SUMMARY_BG)))
        style.append(('VALIGN', (_COL_SUMMARY, 1), (_COL_SUMMARY, 1), 'TOP'))
        style.append(('FONTSIZE', (_COL_SUMMARY, 1), (_COL_SUMMARY, 1), 7.5))

    # Row tinting by result + zebra striping on the leading columns.
    for index, row in enumerate(sheet['rows'], start=1):
        if row['passed']:
            tint = _PASS_BG
        elif row['failedSubjects']:
            tint = _FAIL_BG
        else:
            tint = _REFERRED_BG
        # GPA cell coloured by outcome for a quick scan.
        style.append(('BACKGROUND', (_COL_GPA, index), (_COL_GPA, index),
                      colors.HexColor(tint)))
        if index % 2 == 0:
            style.append(('BACKGROUND', (_COL_SL, index), (_COL_ROLL, index),
                          colors.HexColor(_ALT_ROW)))
        if row['position']:
            style.append(('BACKGROUND', (_COL_POSITION, index), (_COL_POSITION, index),
                          colors.HexColor('#fde68a')))
            style.append(('FONTNAME', (_COL_POSITION, index), (_COL_POSITION, index),
                          'Helvetica-Bold'))

    table.setStyle(TableStyle(style))
    story.append(table)

    footer_style = ParagraphStyle('footer', parent=styles['Normal'], fontSize=7,
                                  textColor=colors.HexColor('#64748b'), spaceBefore=6)
    story.append(Paragraph(
        f"Generated {timezone.localtime():%Y-%m-%d %H:%M} · "
        f"{summary['totalStudent']} students · Sirajganj Polytechnic Institute",
        footer_style,
    ))

    doc.build(story)
    return buffer.getvalue()


# ---------------------------------------------------------------------------
# Personal result card (public portal download)
# ---------------------------------------------------------------------------

def render_student_pdf(payload: dict) -> bytes:
    """A clean A4 result card for one roll, from the public-search payload.

    ``payload`` is the exact dict served by the public search endpoint
    (roll / institute / finalCgpa / results[]), so the PDF can never drift
    from what the portal displays.
    """
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import mm
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.platypus import (
        Paragraph,
        SimpleDocTemplate,
        Spacer,
        Table,
        TableStyle,
    )

    roll = payload['roll']
    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer, pagesize=A4,
        leftMargin=16 * mm, rightMargin=16 * mm,
        topMargin=14 * mm, bottomMargin=14 * mm,
        title=f'BTEB Result — Roll {roll}',
    )
    styles = getSampleStyleSheet()
    brand = ParagraphStyle('brand', parent=styles['Title'], fontSize=16,
                           textColor=colors.HexColor('#047857'), spaceAfter=1)
    subtitle = ParagraphStyle('subtitle', parent=styles['Normal'], fontSize=9.5,
                              alignment=1, textColor=colors.HexColor('#475569'))
    section = ParagraphStyle('section', parent=styles['Heading3'], fontSize=11,
                             textColor=colors.HexColor('#0f172a'), spaceBefore=10,
                             spaceAfter=4)
    body = ParagraphStyle('body', parent=styles['Normal'], fontSize=9, leading=12)
    small = ParagraphStyle('small', parent=body, fontSize=7.5,
                           textColor=colors.HexColor('#64748b'))

    story = [
        Paragraph('BTEB Diploma Result', brand),
        Paragraph('Bangladesh Technical Education Board — semester result transcript', subtitle),
        Spacer(1, 8),
    ]

    institute = payload.get('institute') or {}
    info_rows = [
        ['Roll Number', roll],
        ['Institute', f"{institute.get('code', '')} — {institute.get('name', '')}"
                      if institute else '—'],
    ]
    if payload.get('finalCgpa'):
        info_rows.append(['Final CGPA', payload['finalCgpa']])
    info = Table(info_rows, colWidths=[38 * mm, 140 * mm])
    info.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 9.5),
        ('TEXTCOLOR', (0, 0), (0, -1), colors.HexColor('#334155')),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ('LINEBELOW', (0, 0), (-1, -2), 0.25, colors.HexColor('#e2e8f0')),
    ]))
    story.append(info)

    status_colors = {
        'passed': ('#dcfce7', '#166534', 'Passed'),
        'referred': ('#fee2e2', '#991b1b', 'Referred'),
        'failed': ('#fee2e2', '#991b1b', 'Failed'),
        'expelled': ('#fecaca', '#7f1d1d', 'Expelled'),
        'continuous_fail': ('#ffedd5', '#9a3412', 'Continuous Assessment Failed'),
    }

    for result in payload.get('results', []):
        exam = result['exam']
        bg, fg, label = status_colors.get(
            result['resultType'], ('#f1f5f9', '#334155', result['resultType']),
        )
        story.append(Paragraph(
            f"Semester {exam['semester']} — {exam['regulationYear']} Regulation "
            f"<font size=8 color='#64748b'>({exam['program']})</font>",
            section,
        ))

        meta_bits = [f"Status: <b><font color='{fg}'>{label}</font></b>"]
        if result.get('cgpa'):
            meta_bits.append(f"CGPA: <b>{result['cgpa']}</b>")
        if exam.get('publicationDate'):
            meta_bits.append(f"Published: {exam['publicationDate']}")
        if result.get('expelledRule'):
            meta_bits.append(f"Rule: {result['expelledRule']}")
        story.append(Paragraph(' &nbsp;·&nbsp; '.join(meta_bits), body))

        gpas = sorted(result.get('gpas', []), key=lambda g: g['semester'])
        if gpas:
            header = [f"Sem {g['semester']}" for g in gpas]
            values = ['ref' if g['isReferred'] else str(g['gpa']) for g in gpas]
            gpa_table = Table([header, values],
                              colWidths=[(178 / max(len(gpas), 1)) * mm] * len(gpas))
            table_style = [
                ('GRID', (0, 0), (-1, -1), 0.4, colors.HexColor('#cbd5e1')),
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#f1f5f9')),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 8.5),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('TOPPADDING', (0, 0), (-1, -1), 3),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
            ]
            for col, g in enumerate(gpas):
                if g['isReferred']:
                    table_style.append(
                        ('BACKGROUND', (col, 1), (col, 1), colors.HexColor('#fee2e2')))
                    table_style.append(
                        ('TEXTCOLOR', (col, 1), (col, 1), colors.HexColor('#991b1b')))
            gpa_table.setStyle(TableStyle(table_style))
            story.append(Spacer(1, 3))
            story.append(gpa_table)

        subjects = result.get('subjects', [])
        if subjects:
            rendered = ', '.join(
                s['subjectCode']
                + ('({}{}{})'.format('T' if s['hasTheory'] else '',
                                     ',' if s['hasTheory'] and s['hasPractical'] else '',
                                     'P' if s['hasPractical'] else '')
                   if s['hasTheory'] or s['hasPractical'] else '')
                for s in subjects
            )
            story.append(Spacer(1, 3))
            story.append(Paragraph(f'<b>Subjects to clear:</b> {rendered}', body))

    story.append(Spacer(1, 14))
    story.append(Paragraph(
        f"Generated {timezone.localtime():%d %b %Y, %H:%M} from officially "
        f"published BTEB result notices · verify anytime at "
        f"result.spisg.gov.bd · Sirajganj Polytechnic Institute",
        small,
    ))

    doc.build(story)
    return buffer.getvalue()


# ---------------------------------------------------------------------------
# Excel
# ---------------------------------------------------------------------------

def render_excel(sheet: dict) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = f"{sheet['semesterLabel']} Semester"

    def fill(hex_color: str) -> PatternFill:
        return PatternFill('solid', fgColor=hex_color.lstrip('#'))

    thin = Side(style='thin', color=_BORDER.lstrip('#'))
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left = Alignment(horizontal='left', vertical='center', wrap_text=True)

    headers = HEADERS_XLSX
    # 1-based Excel columns.
    col_summary = _COL_SUMMARY + 1
    col_gpa = _COL_GPA + 1
    col_position = _COL_POSITION + 1
    left_cols = {_COL_NAME + 1, _COL_REFERRED + 1, _COL_FAILED + 1}

    # Title rows.
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    title_cell = ws.cell(row=1, column=1, value=INSTITUTE_NAME)
    title_cell.font = Font(bold=True, size=14, color='1E3A8A')
    title_cell.alignment = center

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(headers))
    sub_cell = ws.cell(row=2, column=1, value=_title(sheet))
    sub_cell.font = Font(size=11, color='334155')
    sub_cell.alignment = center

    header_row = 4
    for col, name in enumerate(headers, start=1):
        c = ws.cell(row=header_row, column=col, value=name)
        c.fill = fill(_HEADER_BG)
        c.font = Font(bold=True, color='FFFFFF', size=10)
        c.alignment = center
        c.border = border

    summary = sheet['summary']
    summary_block = '\n'.join(_summary_lines(summary))

    start = header_row + 1
    for index, row in enumerate(sheet['rows']):
        r = start + index
        tint = _PASS_BG if row['passed'] else (_FAIL_BG if row['failedSubjects'] else _REFERRED_BG)
        # Full summary block goes in the first data row's summary cell (the
        # merge below shows only the top-left cell); newlines + wrap render it
        # as one box.
        values = [
            row['sl'], row['name'], row['gender'], row['roll'], row['gpa'],
            row['referredSubjects'], row['failedSubjects'],
            row['refSubSemWise'], row['position'],
            summary_block if index == 0 else '',
        ]
        for col, value in enumerate(values, start=1):
            c = ws.cell(row=r, column=col, value=value)
            c.border = border
            c.alignment = left if col in left_cols else center
            if col == col_gpa:  # GPA cell tinted by outcome
                c.fill = fill(tint)
            elif col == col_summary:
                c.fill = fill(_SUMMARY_BG)
            elif index % 2 == 1 and col <= _COL_ROLL + 1:
                c.fill = fill(_ALT_ROW)
            if col == col_position and row['position']:
                c.fill = fill('#fde68a')
                c.font = Font(bold=True)

    # Merge the summary column into one block.
    n_rows = len(sheet['rows'])
    if n_rows:
        ws.merge_cells(start_row=start, start_column=col_summary,
                       end_row=start + n_rows - 1, end_column=col_summary)
        ws.cell(row=start, column=col_summary).alignment = Alignment(
            horizontal='left', vertical='top', wrap_text=True,
        )

    #          SL Name Gen Roll GPA Ref Fail RefSub Pos Summary
    widths = [5, 30, 8, 12, 8, 34, 20, 16, 9, 26]
    for col, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col)].width = width

    ws.freeze_panes = ws.cell(row=start, column=1)

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
