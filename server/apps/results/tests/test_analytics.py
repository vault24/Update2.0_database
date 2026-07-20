"""
Analytics endpoint tests: institute/department summaries, comparison data,
CSV download.
"""
from decimal import Decimal

from rest_framework import status
from rest_framework.test import APITestCase

from apps.authentication.models import User
from apps.departments.models import Department
from apps.results.models import (
    Exam,
    Institute,
    ResultImport,
    ResultSubject,
    SemesterGPA,
    StudentResult,
)
from apps.students.models import Student


class AnalyticsApiTests(APITestCase):
    @classmethod
    def setUpTestData(cls):
        cls.cst = Department.objects.create(name='Computer', code='CST')
        cls.eee = Department.objects.create(name='Electrical', code='EEE')
        cls.exam = Exam.objects.create(
            semester=5, regulationYear=2022, program='DIPLOMA IN ENGINEERING',
            heldIn='2025 held in January-March, 2026',
        )
        cls.institute = Institute.objects.create(code='99001', name='Testville')
        cls.import_record = ResultImport.objects.create(
            fileName='5th.pdf', fileSha256='d' * 64, status='completed',
        )

        def result(roll, rtype, gpa5=None, cgpa=None, subjects=()):
            row = StudentResult.objects.create(
                exam=cls.exam, institute=cls.institute,
                importRecord=cls.import_record,
                rollNumber=roll, resultType=rtype, cgpa=cgpa,
            )
            if gpa5 is not None:
                SemesterGPA.objects.create(
                    result=row, semester=5,
                    gpa=None if gpa5 == 'ref' else Decimal(gpa5),
                    isReferred=gpa5 == 'ref',
                )
            for code in subjects:
                ResultSubject.objects.create(result=row, subjectCode=code, hasTheory=True)
            return row

        def student(roll, dept, shift='Morning'):
            return Student.objects.create(
                fullNameEnglish=f'Student {roll}',
                currentRollNumber=roll,
                currentRegistrationNumber=f'REG-{roll}',
                semester=5, department=dept, shift=shift,
            )

        # Results FIRST, then students (auto-sync signal is harmless here).
        result('500001', 'passed', gpa5='3.80')
        result('500002', 'passed', gpa5='3.20')
        result('500003', 'referred', gpa5='ref', subjects=['28551', '26811'])
        result('500004', 'failed', subjects=['28551'])
        result('999001', 'passed', gpa5='3.99')  # national-only (no student)

        student('500001', cls.cst, 'Morning')
        student('500002', cls.cst, 'Day')
        student('500003', cls.eee, 'Morning')
        student('500004', cls.eee, 'Morning')

        cls.admin = User.objects.create_user(
            username='analyticsadmin', email='aa@x.com', password='pw',
            role='registrar', account_status='active',
        )

    def setUp(self):
        self.client.force_authenticate(self.admin)

    def test_semesters_list(self):
        response = self.client.get('/api/results/analytics/semesters/')
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        data = response.json()
        sem5 = next(s for s in data if s['semester'] == 5)
        self.assertEqual(sem5['students'], 4)  # our 4 matched students
        self.assertEqual(sem5['label'], '5th Semester')

    def test_summary(self):
        response = self.client.get('/api/results/analytics/summary/?semester=5')
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        data = response.json()

        institute = data['institute']
        self.assertEqual(institute['appeared'], 4)
        self.assertEqual(institute['passed'], 2)
        self.assertEqual(institute['referred'], 1)
        self.assertEqual(institute['failed'], 1)
        self.assertEqual(institute['passRate'], 50.0)
        self.assertEqual(institute['avgGpa'], 3.5)  # (3.80 + 3.20) / 2

        departments = {d['name']: d for d in data['departments']}
        self.assertEqual(departments['Computer']['passed'], 2)
        self.assertEqual(departments['Computer']['passRate'], 100.0)
        self.assertEqual(departments['Electrical']['passed'], 0)
        self.assertIn('Morning', departments['Electrical']['shifts'])

        self.assertEqual(data['topFailedSubjects'][0]['subjectCode'], '28551')
        self.assertEqual(data['topFailedSubjects'][0]['students'], 2)

        self.assertEqual(data['topPerformers'][0]['roll'], '500001')

        self.assertEqual(data['national']['records'], 5)
        self.assertEqual(data['national']['passed'], 3)

    def test_summary_requires_semester(self):
        response = self.client.get('/api/results/analytics/summary/')
        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)

    def test_pdf_download(self):
        response = self.client.get('/api/results/analytics/download/?semester=5')
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(response['Content-Type'], 'application/pdf')
        self.assertTrue(response.content.startswith(b'%PDF'))
        self.assertIn('.pdf', response['Content-Disposition'])

    def test_excel_download(self):
        response = self.client.get(
            '/api/results/analytics/download/?semester=5&type=excel'
        )
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertIn('spreadsheetml', response['Content-Type'])
        # xlsx is a zip: starts with PK.
        self.assertTrue(response.content.startswith(b'PK'))

    def test_download_department_shift_filter(self):
        # Excel makes the row count easy to inspect via openpyxl.
        from io import BytesIO

        from openpyxl import load_workbook

        response = self.client.get(
            f'/api/results/analytics/download/?semester=5'
            f'&department={self.eee.id}&shift=Morning&type=excel'
        )
        wb = load_workbook(BytesIO(response.content))
        ws = wb.active
        names = [
            ws.cell(row=r, column=2).value
            for r in range(5, ws.max_row + 1)
            if ws.cell(row=r, column=2).value
        ]
        self.assertIn('Student 500003', names)
        self.assertIn('Student 500004', names)
        self.assertNotIn('Student 500001', names)

    def test_analytics_denied_for_students(self):
        student_user = User.objects.create_user(
            username='plainstudent', email='pl@x.com', password='pw',
            role='student', account_status='active',
        )
        self.client.force_authenticate(student_user)
        for url in (
            '/api/results/analytics/semesters/',
            '/api/results/analytics/summary/?semester=5',
            '/api/results/analytics/download/?semester=5',
        ):
            self.assertEqual(
                self.client.get(url).status_code, status.HTTP_403_FORBIDDEN, url,
            )
